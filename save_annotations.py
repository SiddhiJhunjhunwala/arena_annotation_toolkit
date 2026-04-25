"""
Step 3: Export annotations to Excel.

Run manually: python save_annotations.py
Or via the UI: click "Export Excel" button.

Output: data/arena_annotations.xlsx  (2 sheets: Annotations, Summary)
"""

import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = Path("data/arena_candidates.json")
OUT_FILE  = Path("data/arena_annotations.xlsx")

def thin_border():
    s = Side(style="thin", color="252A35")
    return Border(left=s, right=s, top=s, bottom=s)

def export():
    with open(DATA_FILE) as f:
        records = json.load(f)

    rows = []
    for i, rec in enumerate(records):
        ann = rec.get("annotation", {})

        rows.append({
            # ── Your requested columns ─────────────────────────────────────
            "Conv ID":              rec.get("conv_id") or rec.get("question_id", ""),
            "Turn":                 rec.get("turn", ""),
            "Feedback Type":        rec.get("feedback_type", ""),
            "How It's Given":       rec.get("how_its_given", ""),
            "Context":              rec.get("context", ""),
            "What's Updated":       ann.get("what_is_updated", ""),
            "Notes":                ann.get("notes", ""),
            "User Message Preview": rec.get("user_msg_preview", ""),
            "Task Domain":          ann.get("task_domain", ""),
            # ── Extra useful columns ───────────────────────────────────────
            "Confirmed Signal":     ann.get("confirmed_signal", ""),
            "Confidence":           ann.get("confidence", ""),
            "Signal Evidence":      ann.get("signal_evidence", ""),
            "Inferred Preference":  ann.get("inferred_preference", ""),
            "Winner":               rec.get("winner", ""),
            "Model A":              rec.get("model_a", ""),
            "Model B":              rec.get("model_b", ""),
            "Num User Turns":       rec.get("num_user_turns", ""),
            "All Detected Signals": ", ".join(rec.get("detected_signals", [])),
            "Annotated":            1 if ann.get("confirmed_signal") is not None else 0,
        })

    df = pd.DataFrame(rows)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    annotated = df[df["Annotated"] == 1]
    summary_rows = []
    for col in ["Confirmed Signal", "Confidence", "Task Domain"]:
        for val, cnt in annotated[col].value_counts().items():
            summary_rows.append({"Category": col, "Value": val, "Count": int(cnt)})
    if not annotated.empty:
        ct = pd.crosstab(annotated["Confirmed Signal"], annotated["Task Domain"])
        for sig in ct.index:
            for dom in ct.columns:
                summary_rows.append({
                    "Category": "Signal × Domain",
                    "Value":    f"{sig} × {dom}",
                    "Count":    int(ct.loc[sig, dom]),
                })
    df_summary = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(columns=["Category","Value","Count"])

    # ── Write ──────────────────────────────────────────────────────────────────
    with pd.ExcelWriter(str(OUT_FILE), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Annotations", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    # ── Styling ────────────────────────────────────────────────────────────────
    wb = load_workbook(str(OUT_FILE))

    DARK_BG  = "0D0F14"
    HDR_BG   = "151820"
    ACCENT   = "4FFFB0"
    ACCENT2  = "FF6B6B"
    ACCENT3  = "FFCC44"
    ACCENT4  = "74B9FF"
    TEXT     = "E2E8F0"
    MUTED    = "64748B"

    SIG_COLORS = {
        "response_ignoring":  ACCENT4,
        "frustration_marker": ACCENT2,
        "task_abandonment":   ACCENT3,
        "none":               MUTED,
    }

    # Column widths for Annotations sheet (match column order in df)
    COL_WIDTHS = {
        "Conv ID":              18,
        "Turn":                  8,
        "Feedback Type":        22,
        "How It's Given":       38,
        "Context":              45,
        "What's Updated":       35,
        "Notes":                28,
        "User Message Preview": 45,
        "Task Domain":          18,
        "Confirmed Signal":     22,
        "Confidence":           12,
        "Signal Evidence":      30,
        "Inferred Preference":  40,
        "Winner":               12,
        "Model A":              18,
        "Model B":              18,
        "Num User Turns":       14,
        "All Detected Signals": 28,
        "Annotated":            10,
    }

    WRAP_COLS = {"How It's Given", "Context", "What's Updated", "Notes",
                 "User Message Preview", "Signal Evidence", "Inferred Preference"}

    def style_annotations(ws):
        # Header
        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor=HDR_BG)
            cell.font      = Font(name="Arial", bold=True, color=ACCENT, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border()

        # Column widths
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = COL_WIDTHS.get(header, 16)

        fill_dark = PatternFill("solid", fgColor=DARK_BG)
        fill_alt  = PatternFill("solid", fgColor=HDR_BG)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = fill_dark if row_idx % 2 == 0 else fill_alt

            # Find confirmed signal value for color
            confirmed_col = headers.index("Confirmed Signal") if "Confirmed Signal" in headers else -1
            sig_val = row[confirmed_col].value if confirmed_col >= 0 else ""
            sig_color = SIG_COLORS.get(sig_val or "", "")

            for col_idx, cell in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else ""
                cell.fill   = fill
                cell.border = thin_border()
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(header in WRAP_COLS),
                )
                if header == "Confirmed Signal" and sig_color:
                    cell.font = Font(name="Arial", color=sig_color, bold=True, size=10)
                elif header in WRAP_COLS:
                    cell.font = Font(name="Arial", color=TEXT, size=10, italic=True)
                elif header == "Turn":
                    cell.font = Font(name="Arial", color=ACCENT3, bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.font = Font(name="Arial", color=TEXT, size=10)

            ws.row_dimensions[row_idx].height = 55

        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

    def style_summary(ws):
        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor=HDR_BG)
            cell.font      = Font(name="Arial", bold=True, color=ACCENT, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 12
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor=DARK_BG if row_idx % 2 == 0 else HDR_BG)
            for cell in row:
                cell.fill      = fill
                cell.font      = Font(name="Arial", color=TEXT, size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border()
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

    style_annotations(wb["Annotations"])
    style_summary(wb["Summary"])
    wb.save(str(OUT_FILE))

    total     = len(records)
    done      = sum(1 for r in records if r["annotation"]["confirmed_signal"] is not None)
    print(f"Exported {total} records ({done} annotated) → {OUT_FILE}")
    return str(OUT_FILE)


if __name__ == "__main__":
    export()